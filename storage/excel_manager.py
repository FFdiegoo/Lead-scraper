"""Beheert leads.xlsx: aanmaken, lezen, schrijven, status bijwerken."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import config

COLUMNS = [
    "Bedrijf",
    "Adres",
    "Telefoon",
    "Website",
    "Maps URL",
    "Categorie",
    "Status",
    "Naam Eigenaar",
    "Email Eigenaar",
    "Notities",
    "Toegevoegd Op",
    "Email Verstuurd Op",
]

COL = {name: idx + 1 for idx, name in enumerate(COLUMNS)}

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
COL_WIDTHS = [32, 38, 16, 34, 50, 22, 16, 26, 34, 38, 18, 20]


def get_workbook() -> openpyxl.Workbook:
    path = config.LEADS_FILE
    if path.exists():
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    _write_header(ws)
    wb.save(path)
    return wb


def _write_header(ws) -> None:
    bold_white = Font(bold=True, color=HEADER_FG)
    fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = bold_white
        cell.fill = fill
        cell.alignment = center

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def get_existing_companies(wb: openpyxl.Workbook) -> set:
    ws = wb.active
    return {
        str(row[COL["Bedrijf"] - 1]).lower().strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[COL["Bedrijf"] - 1]
    }


def add_lead(wb: openpyxl.Workbook, lead: dict) -> None:
    ws = wb.active
    row = ws.max_row + 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws.cell(row=row, column=COL["Bedrijf"], value=lead.get("name", ""))
    ws.cell(row=row, column=COL["Adres"], value=lead.get("address", ""))
    ws.cell(row=row, column=COL["Telefoon"], value=lead.get("phone", ""))
    ws.cell(row=row, column=COL["Website"], value=lead.get("website", ""))
    ws.cell(row=row, column=COL["Maps URL"], value=lead.get("maps_url", ""))
    ws.cell(row=row, column=COL["Categorie"], value=lead.get("category", ""))
    ws.cell(row=row, column=COL["Status"], value="Nieuw")
    ws.cell(row=row, column=COL["Naam Eigenaar"], value=lead.get("owner_name", ""))
    ws.cell(row=row, column=COL["Email Eigenaar"], value=lead.get("owner_email", ""))
    ws.cell(row=row, column=COL["Notities"], value=lead.get("notes", ""))
    ws.cell(row=row, column=COL["Toegevoegd Op"], value=now)
    ws.cell(row=row, column=COL["Email Verstuurd Op"], value="")


def get_pending_emails(wb: openpyxl.Workbook) -> list:
    """Geef leads terug die 48u+ oud zijn, status Nieuw hebben en een email adres hebben."""
    ws = wb.active
    now = datetime.now()
    pending = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        status = row[COL["Status"] - 1]
        owner_email = row[COL["Email Eigenaar"] - 1]
        added_on_str = row[COL["Toegevoegd Op"] - 1]

        owner_name = row[COL["Naam Eigenaar"] - 1] or ""

        if status != "Nieuw" or not owner_email or not added_on_str:
            continue

        # Sla over als voornaam ontbreekt
        if not owner_name or not owner_name.strip():
            continue

        # Controleer of de voornaam voorkomt in het emailadres (local part voor @)
        first_name = owner_name.strip().split()[0].lower()
        email_local = owner_email.split("@")[0].lower()
        if first_name not in email_local and first_name[:1] not in email_local.split(".")[0]:
            continue

        try:
            added_on = datetime.strptime(str(added_on_str), "%Y-%m-%d %H:%M")
            hours_passed = (now - added_on).total_seconds() / 3600
            if hours_passed >= config.EMAIL_DELAY_HOURS:
                pending.append({
                    "row": row_idx,
                    "name": row[COL["Bedrijf"] - 1],
                    "owner_name": owner_name,
                    "owner_email": owner_email,
                })
        except ValueError:
            pass

    return pending


def mark_email_sent(wb: openpyxl.Workbook, row_idx: int) -> None:
    ws = wb.active
    ws.cell(row=row_idx, column=COL["Status"], value="Email Verstuurd")
    ws.cell(
        row=row_idx,
        column=COL["Email Verstuurd Op"],
        value=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )


def save(wb: openpyxl.Workbook) -> None:
    wb.save(config.LEADS_FILE)
